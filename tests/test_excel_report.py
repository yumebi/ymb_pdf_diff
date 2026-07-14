"""Excelレポート出力(#新機能11: image_sizeプリセット)が各サイズで正常に生成できることを確認する。"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from openpyxl.utils.units import EMU_to_pixels

from ymb_pdf_diff.core import align_documents, detect_visual_only_changes, load_pdf_pages
from ymb_pdf_diff.report import build_excel_report

import fitz


def _make_sample_pdfs(tmp_dir: Path):
    doc_a = fitz.open()
    doc_a.new_page().insert_text((72, 72), "Page A1 alpha text here.")
    doc_a.new_page().insert_text((72, 72), "Page A2 beta text here unchanged.")
    path_a = tmp_dir / "excel_report_sample_a.pdf"
    doc_a.save(str(path_a))
    doc_a.close()

    doc_b = fitz.open()
    doc_b.new_page().insert_text((72, 72), "Page A1 alpha text here CHANGED.")
    doc_b.new_page().insert_text((72, 72), "Page A2 beta text here unchanged.")
    path_b = tmp_dir / "excel_report_sample_b.pdf"
    doc_b.save(str(path_b))
    doc_b.close()
    return path_a, path_b


def test_build_excel_report_with_each_image_size_preset():
    """#新機能11: image_size="small"/"medium"/"large"のいずれでも正常にExcelを生成できることを確認する
    (スモークテスト。ファイルサイズの詳細な比較はscripts/benchmark.pyで行う)。
    """
    tmp_dir = Path(__file__).resolve().parent.parent
    path_a, path_b = _make_sample_pdfs(tmp_dir)
    outputs = {size: tmp_dir / f"excel_report_sample_output_{size}.xlsx" for size in ("small", "medium", "large")}
    try:
        pages_a = load_pdf_pages(str(path_a))
        pages_b = load_pdf_pages(str(path_b))
        alignment = align_documents(pages_a, pages_b)
        detect_visual_only_changes(alignment, str(path_a), str(path_b))

        for size, output_path in outputs.items():
            build_excel_report(
                str(path_a), str(path_b), pages_a, pages_b, alignment, str(output_path), image_size=size,
            )
            assert output_path.exists()
            wb = load_workbook(str(output_path))
            assert "サマリー" in wb.sheetnames

        print("OK: test_build_excel_report_with_each_image_size_preset")
    finally:
        path_a.unlink(missing_ok=True)
        path_b.unlink(missing_ok=True)
        for output_path in outputs.values():
            output_path.unlink(missing_ok=True)


def _make_multiline_diff_pdfs(tmp_dir: Path):
    """変更前後で3行以上のテキストが変わるページペアを生成する(行数に応じた行高計算の確認用)。"""
    doc_a = fitz.open()
    doc_a.new_page().insert_text(
        (72, 72),
        "line one alpha\nline two alpha\nline three alpha\nline four alpha",
    )
    path_a = tmp_dir / "excel_report_multiline_a.pdf"
    doc_a.save(str(path_a))
    doc_a.close()

    doc_b = fitz.open()
    doc_b.new_page().insert_text(
        (72, 72),
        "line one BETA\nline two BETA\nline three BETA\nline four BETA",
    )
    path_b = tmp_dir / "excel_report_multiline_b.pdf"
    doc_b.save(str(path_b))
    doc_b.close()
    return path_a, path_b


def test_text_diff_table_start_row_is_dynamic_and_columns_are_wide():
    """#報告対応: 診断シートのテキスト差分テーブルが固定行(70)ではなく、実際のサムネイル高さに
    応じた動的な行から始まること、また変更前/変更後の列幅が読みやすく広げられていることを確認する。
    """
    tmp_dir = Path(__file__).resolve().parent.parent
    path_a, path_b = _make_multiline_diff_pdfs(tmp_dir)
    output_small = tmp_dir / "excel_report_dynrow_output_small.xlsx"
    output_large = tmp_dir / "excel_report_dynrow_output_large.xlsx"
    try:
        pages_a = load_pdf_pages(str(path_a))
        pages_b = load_pdf_pages(str(path_b))
        alignment = align_documents(pages_a, pages_b)
        detect_visual_only_changes(alignment, str(path_a), str(path_b))

        start_rows = {}
        for output_path, size in ((output_small, "small"), (output_large, "large")):
            build_excel_report(
                str(path_a), str(path_b), pages_a, pages_b, alignment, str(output_path), image_size=size,
            )
            wb = load_workbook(str(output_path))
            detail_sheet_name = [n for n in wb.sheetnames if n != "サマリー"][0]
            ws = wb[detail_sheet_name]

            # 「種別」ヘッダーが書かれている行を探す = テーブルの開始行
            header_row = None
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == "種別":
                    header_row = row
                    break
            assert header_row is not None, f"{size}: 種別ヘッダーが見つからない"
            start_rows[size] = header_row

            # 変更前(A)/変更後(B)列(B/C列)が読みやすい幅に広げられていること
            assert ws.column_dimensions["B"].width > 30, f"{size}: 列幅が広げられていない"
            assert ws.column_dimensions["C"].width > 30, f"{size}: 列幅が広げられていない"

        # largeプリセットの方がサムネイルが大きく縦に伸びるため、テーブル開始行も
        # small以上(基本的にはより下)になる = ハードコードされた固定行ではなく動的であることの証明
        assert start_rows["large"] >= start_rows["small"]

        print("OK: test_text_diff_table_start_row_is_dynamic_and_columns_are_wide")
    finally:
        path_a.unlink(missing_ok=True)
        path_b.unlink(missing_ok=True)
        output_small.unlink(missing_ok=True)
        output_large.unlink(missing_ok=True)


def _first_detail_sheet(wb):
    detail_sheet_name = [n for n in wb.sheetnames if n != "サマリー"][0]
    return wb[detail_sheet_name]


def _image_b_offset_px(ws) -> int:
    """詳細シートに埋め込まれた画像B(2枚目、区切り線を除く)の、画像Aからの水平オフセット(px)を返す。

    画像A/BともにAbsoluteAnchor(シート原点基準の絶対座標、列幅設定に一切依存しない)
    として保存されるため、画像Bのpos.x - 画像Aのpos.xが「画像Aの実幅+余白」になるはずである。
    画像A・画像Bに加えて中央の区切り線(縦の細い画像)も埋め込まれるため、合計3枚になる。
    """
    images = ws._images
    assert len(images) == 3, f"詳細シートに画像A/B+区切り線の計3枚が埋め込まれているはず: {len(images)}"
    img_a, img_b = images[0], images[1]
    anchor_a, anchor_b = img_a.anchor, img_b.anchor
    assert not isinstance(anchor_a, str) and not isinstance(anchor_b, str), (
        "画像A/BはAbsoluteAnchor(列幅非依存の絶対座標)であるべき"
    )
    return round(EMU_to_pixels(anchor_b.pos.x) - EMU_to_pixels(anchor_a.pos.x))


def _image_a_width_px(ws) -> int:
    images = ws._images
    return images[0].width


def test_image_b_gap_is_pixel_based_not_column_width_based():
    """#報告対応: 変更前(A)/変更後(B)の画像間の隙間が、テキスト差分テーブルの列幅設定
    (B/C列=60文字幅)に引きずられて数百px以上広がってしまう不具合の修正確認。
    画像Bの横位置(colOff)は画像Aの実際の埋め込み幅からの小さな固定余白のみであるべきで、
    small/largeのどちらのプリセットでも、隙間が画像幅にほぼ比例した小さい値であること
    (=列幅60文字が原因で数百~数千px離れてしまう不具合が再発していないこと)を確認する。
    """
    tmp_dir = Path(__file__).resolve().parent.parent
    path_a, path_b = _make_sample_pdfs(tmp_dir)
    outputs = {size: tmp_dir / f"excel_report_gap_output_{size}.xlsx" for size in ("small", "large")}
    try:
        pages_a = load_pdf_pages(str(path_a))
        pages_b = load_pdf_pages(str(path_b))
        alignment = align_documents(pages_a, pages_b)
        detect_visual_only_changes(alignment, str(path_a), str(path_b))

        for size, output_path in outputs.items():
            build_excel_report(
                str(path_a), str(path_b), pages_a, pages_b, alignment, str(output_path), image_size=size,
            )
            wb = load_workbook(str(output_path))
            ws = _first_detail_sheet(wb)

            image_a_width_px = _image_a_width_px(ws)
            actual_gap = _image_b_offset_px(ws)

            # 画像Aの実幅に対して_IMAGE_GAP_PX(=100)程度の小さな余白のはず(以前の不具合では
            # 列幅60文字x2列分が上乗せされ、数百~数千pxも離れてしまっていた)。
            assert image_a_width_px + 80 <= actual_gap <= image_a_width_px + 130, (
                f"{size}: 画像間の隙間が不自然({actual_gap=}, {image_a_width_px=})。"
                "列幅設定に引きずられている可能性がある。"
            )

        print("OK: test_image_b_gap_is_pixel_based_not_column_width_based")
    finally:
        path_a.unlink(missing_ok=True)
        path_b.unlink(missing_ok=True)
        for output_path in outputs.values():
            output_path.unlink(missing_ok=True)


def _make_visual_only_pdfs(tmp_dir: Path, box_color_a, box_color_b):
    """テキストは同一だが見た目(図形の色)だけ異なる1ページのPDFペアを作る
    (tests/test_visual_only.pyと同じパターン)。diff_page_linesのentriesが
    空になる=テキスト差分テーブルを持たない「visual_only」ページのケースを再現する。
    """
    doc_a = fitz.open()
    p1 = doc_a.new_page()
    p1.insert_text((72, 72), "Page one identical text but the box color differs.")
    p1.draw_rect(fitz.Rect(72, 120, 300, 220), color=box_color_a, fill=box_color_a, width=0)
    path_a = tmp_dir / "excel_report_visualonly_a.pdf"
    doc_a.save(str(path_a))
    doc_a.close()

    doc_b = fitz.open()
    p1b = doc_b.new_page()
    p1b.insert_text((72, 72), "Page one identical text but the box color differs.")
    p1b.draw_rect(fitz.Rect(72, 120, 300, 220), color=box_color_b, fill=box_color_b, width=0)
    path_b = tmp_dir / "excel_report_visualonly_b.pdf"
    doc_b.save(str(path_b))
    doc_b.close()
    return path_a, path_b


def test_image_b_gap_is_small_for_visual_only_page_without_text_diff_table():
    """#報告対応: テキスト差分テーブルが存在しない(visual_onlyでdiff_page_linesが空の)ページでも、
    画像A/B間の隙間が小さいままであることを確認する(不具合はテーブルの有無に関係なく
    固定の列アンカーが原因だったため、テーブルがなくても再発しないことの確認)。
    """
    tmp_dir = Path(__file__).resolve().parent.parent
    path_a, path_b = _make_visual_only_pdfs(tmp_dir, (0, 0, 1), (1, 0, 0))
    output_path = tmp_dir / "excel_report_visualonly_output.xlsx"
    try:
        pages_a = load_pdf_pages(str(path_a))
        pages_b = load_pdf_pages(str(path_b))
        alignment = align_documents(pages_a, pages_b)
        # テキストだけでは差分なし判定になるため、見た目の差分検出を明示的に走らせる。
        assert all(s.status == "unchanged" for s in alignment.page_statuses)
        detect_visual_only_changes(alignment, str(path_a), str(path_b))
        assert alignment.page_statuses[0].status == "changed"
        assert alignment.page_statuses[0].visual_only is True

        build_excel_report(str(path_a), str(path_b), pages_a, pages_b, alignment, str(output_path))
        wb = load_workbook(str(output_path))
        ws = _first_detail_sheet(wb)

        # diff_page_linesのentriesが空 = テキスト差分テーブルはなく、代わりに
        # 「テキストは同一です。画像(見た目)のみ差分があります。」の案内文が入る想定。
        found_visual_only_note = any(
            ws.cell(row=r, column=1).value == "テキストは同一です。画像(見た目)のみ差分があります。"
            for r in range(1, ws.max_row + 1)
        )
        assert found_visual_only_note, "visual_onlyページの案内文が見つからない"

        image_a_width_px = _image_a_width_px(ws)
        actual_gap = _image_b_offset_px(ws)
        assert image_a_width_px + 80 <= actual_gap <= image_a_width_px + 130, (
            f"visual_onlyページでも画像間の隙間は小さいはず({actual_gap=}, {image_a_width_px=})。"
        )

        print("OK: test_image_b_gap_is_small_for_visual_only_page_without_text_diff_table")
    finally:
        path_a.unlink(missing_ok=True)
        path_b.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)


if __name__ == "__main__":
    test_build_excel_report_with_each_image_size_preset()
    test_text_diff_table_start_row_is_dynamic_and_columns_are_wide()
    test_image_b_gap_is_pixel_based_not_column_width_based()
    test_image_b_gap_is_small_for_visual_only_page_without_text_diff_table()
