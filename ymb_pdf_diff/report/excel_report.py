import io
import math
import re
from datetime import datetime
from typing import Callable, Dict, List, Optional, Set

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageDraw

from ..core import (
    AlignmentResult,
    PageLine,
    PageStatus,
    diff_page_lines,
    diff_page_pair,
    draw_highlights,
    render_page,
)
from .image_size import DEFAULT_IMAGE_SIZE, resolve_excel_thumb_max_size

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

_STATUS_LABEL = {
    "unchanged": "差分なし",
    "changed": "差分あり",
    "inserted": "追加ページ(Bのみ)",
    "deleted": "削除ページ(Aのみ)",
}

_FILL_DELETE = PatternFill("solid", fgColor="FFC7CE")
_FONT_DELETE = Font(color="9C0006")
_FILL_INSERT = PatternFill("solid", fgColor="C6EFCE")
_FONT_INSERT = Font(color="006100")
_FILL_REPLACE = PatternFill("solid", fgColor="FFEB9C")
_FONT_REPLACE = Font(color="9C6500")
_HEADER_FONT = Font(bold=True)
_JPEG_QUALITY = 85  # 埋め込みJPEGの画質(#新機能11)。PNGよりはるかに小さく、見た目の劣化も目立たない

_DETAIL_IMAGE_ANCHOR_ROW = 5  # 変更前/変更後サムネイルを埋め込む行
_EXCEL_ROW_HEIGHT_PX = 20  # 既定の行高(15pt)をピクセル換算した近似値(96DPIでは1pt≒1.333px)
_MIN_TEXT_DIFF_START_ROW = 20  # サムネイルが小さい場合でも窮屈にならないよう最低限確保する開始行
_TEXT_DIFF_COLUMN_WIDTH = 60  # 変更前(A)/変更後(B)列の幅(文字単位)。読みやすさのため広めに確保

# #報告対応: 変更前(A)/変更後(B)の画像間の隙間がテキスト差分テーブルの列幅設定
# (B/C列=60文字幅)に引きずられて数百px以上に広がってしまう不具合の修正用定数。
# 画像Bの横位置は列幅に一切依存させず、画像Aの実ピクセル幅からの固定オフセットで決める。
# 40pxでは画面によってはまだ「かぶって見える」との指摘があったため、余白と枠線を
# 大幅に強化し、誰の画面でも一目で別画像だと分かるようにする。
_IMAGE_GAP_PX = 100  # 画像Aと画像Bの間に空ける余白(px)。中央に区切り線も描くため広めに確保
_PX_PER_CHAR_UNIT = 7  # Excel列幅の「文字単位」を近似的にpx換算する係数
_DEFAULT_COL_WIDTH_PX = 64  # 既定の列幅(約8.43文字)をpx換算した近似値

# #報告対応: ページの背景が白いことが多く、画像自体に枠がないと隣接する画像A/Bの境界が
# 見分けられず「かぶって(重なって)いるように見える」問題への対応。JPEG化する前に
# サムネイルの外周へ濃いグレーの太い枠線を描き、各画像の実際の範囲を常に視認できるようにする。
_THUMBNAIL_BORDER_COLOR = (90, 90, 90)
_THUMBNAIL_BORDER_WIDTH = 4
# 画像A/Bの隙間の中央に描く縦の区切り線(#報告対応)。余白があるだけでは離れて見えにくい
# という指摘に対応し、「ここが境目」と一目で分かる明示的な仕切りを追加する。
_DIVIDER_COLOR_HEX = "#808080"
_DIVIDER_WIDTH_PX = 3


def _safe_sheet_name(name: str, used: Set[str]) -> str:
    name = _INVALID_SHEET_CHARS.sub("_", name)[:31]
    candidate = name
    suffix = 1
    while candidate in used:
        suffix += 1
        candidate = f"{name[: 31 - len(str(suffix)) - 1]}_{suffix}"
    used.add(candidate)
    return candidate


def _sheet_name_for(status: PageStatus, used: Set[str]) -> str:
    if status.status in ("changed",) and status.a_page is not None and status.b_page is not None:
        base = f"P_A{status.a_page + 1}_B{status.b_page + 1}"
    elif status.status == "inserted":
        base = f"P_B{status.b_page + 1}_new"
    elif status.status == "deleted":
        base = f"P_A{status.a_page + 1}_del"
    else:
        base = f"P_A{status.a_page + 1}"
    return _safe_sheet_name(base, used)


def _diff_count_for(status: PageStatus, pages_a: List[List[PageLine]], pages_b: List[List[PageLine]]) -> int:
    if status.status == "changed" and status.a_page is not None and status.b_page is not None:
        return len(diff_page_lines(pages_a[status.a_page], pages_b[status.b_page]))
    if status.status in ("inserted", "deleted"):
        return 1
    return 0


def _absolute_position_emu(anchor_row: int, x_offset_px: int) -> XDRPoint2D:
    """行番号(1-based)とA1セル基準のピクセルオフセットから、シート原点基準の絶対EMU座標を作る。

    #報告対応: 当初はOneCellAnchor + AnchorMarker.colOff(列起点からのピクセルオフセット)で
    列幅非依存の配置を狙ったが、実機Excelで検証したところ「colOffが起点列自体の実際の幅を
    超えると、意図した位置まで描画されず隣の画像とくっつく/被る」という実挙動が確認された
    (列幅60文字でもcolOff=900pxのような大きな値は列幅を超えるため破綻していた)。
    これを避けるため、セル相対のOneCellAnchorをやめ、シート全体の原点(A1の左上)からの
    絶対座標で配置するAbsoluteAnchorに切り替える。これなら列幅の解釈に一切依存しない。
    A列の幅は明示的に設定していないため既定幅(_DEFAULT_COL_WIDTH_PX)とみなし、
    そこからx_offset_px分右に、行はデフォルト行高(_EXCEL_ROW_HEIGHT_PX)の積算で下にずらす。
    """
    x_px = _DEFAULT_COL_WIDTH_PX + x_offset_px
    y_px = (anchor_row - 1) * _EXCEL_ROW_HEIGHT_PX
    return XDRPoint2D(pixels_to_EMU(x_px), pixels_to_EMU(y_px))


def _embed_thumbnail(ws: Worksheet, image, anchor_row: int, thumb_max_size, x_offset_px: int = 0) -> tuple:
    """PIL Imageをサムネイル化してExcelに埋め込み、実際のサムネイルサイズ(幅, 高さ)を返す。

    openpyxlのImage._data()は`fp`(元ファイルのファイルポインタ)を読みに行く実装のため、
    PIL.Image.new/copy等で生成した(ファイルから開いていない)画像はそのまま渡すと
    `AttributeError: 'Image' object has no attribute 'fp'`になる。JPEGとしてメモリに
    一度書き出し(#新機能11でPNGから変更、同じサイズでファイルがはるかに小さくなる)、
    fp/formatを明示的に持たせてから渡す。openpyxlはfp/formatを見るだけなのでJPEGでも同様に動作する。

    戻り値のサイズは、後段のテキスト差分テーブルの開始行を実際の埋め込み高さに応じて
    動的に計算するために使う(画像サイズプリセットによって縦横比が変わるため)。

    x_offset_px(#報告対応)はA1セル基準のピクセルオフセットで、AbsoluteAnchorにより
    列幅設定から完全に独立して配置される(詳細は_absolute_position_emuのdocstring参照)。
    """
    thumb = image.copy()
    thumb.thumbnail(thumb_max_size)
    if thumb.mode != "RGB":
        thumb = thumb.convert("RGB")
    # ページ背景が白いことが多く、枠がないと隣接する画像同士の境界が視認できず
    # 「重なっている」ように見えてしまうため、サムネイル全体を薄いグレーで縁取る。
    ImageDraw.Draw(thumb).rectangle(
        [0, 0, thumb.width - 1, thumb.height - 1],
        outline=_THUMBNAIL_BORDER_COLOR,
        width=_THUMBNAIL_BORDER_WIDTH,
    )
    buf = io.BytesIO()
    thumb.save(buf, format="JPEG", quality=_JPEG_QUALITY)
    buf.seek(0)
    thumb.fp = buf
    thumb.format = "JPEG"
    xl_image = XLImage(thumb)
    xl_image.anchor = AbsoluteAnchor(
        pos=_absolute_position_emu(anchor_row, x_offset_px),
        ext=XDRPositiveSize2D(pixels_to_EMU(thumb.size[0]), pixels_to_EMU(thumb.size[1])),
    )
    ws.add_image(xl_image)
    return thumb.size


def _embed_divider(ws: Worksheet, anchor_row: int, x_offset_px: int, height_px: int) -> None:
    """画像A/Bの隙間の中央に、縦の区切り線を1本描く(#報告対応)。

    余白を空けるだけでは「離れているのか、たまたま隙間が空いているだけなのか」が
    伝わりにくいという指摘への対応。明示的な仕切り線を入れることで、ここが
    2つの画像の境目であることを一目で示す。_embed_thumbnailと同じ
    AbsoluteAnchorの仕組みを使い、列幅設定に依存せず正確な位置に配置する。
    """
    divider = Image.new("RGB", (max(_DIVIDER_WIDTH_PX, 1), max(height_px, 1)), _DIVIDER_COLOR_HEX)
    buf = io.BytesIO()
    divider.save(buf, format="PNG")
    buf.seek(0)
    divider.fp = buf
    divider.format = "PNG"
    xl_image = XLImage(divider)
    xl_image.anchor = AbsoluteAnchor(
        pos=_absolute_position_emu(anchor_row, x_offset_px),
        ext=XDRPositiveSize2D(pixels_to_EMU(divider.width), pixels_to_EMU(divider.height)),
    )
    ws.add_image(xl_image)


def _approximate_column_for_pixel_offset(
    offset_px: int,
    widened_cols: int = 2,
    widened_col_px: float = _TEXT_DIFF_COLUMN_WIDTH * _PX_PER_CHAR_UNIT,
    default_col_px: float = _DEFAULT_COL_WIDTH_PX,
) -> int:
    """列Bの左端からのピクセルオフセットが、おおよそどの列(1-based)に位置するかを概算する。

    これは「変更後(B)」という見出しラベルの表示位置を決めるためだけの概算であり、
    画像本体の配置(_embed_thumbnailのAbsoluteAnchorによりピクセル精度で決まる)
    には一切影響しない。_write_text_diff_tableでB/C列は_TEXT_DIFF_COLUMN_WIDTH(60文字)
    に固定される想定のため、その幅を1文字あたり約_PX_PER_CHAR_UNIT pxとして概算し、
    それ以降の列は既定幅(約_DEFAULT_COL_WIDTH_PX px)とみなして計算する。
    ラベルは装飾的なテキストなので、半列程度のズレは許容する簡易計算とする。
    """
    remaining_px = max(0, offset_px - widened_cols * widened_col_px)
    extra_cols = math.ceil(remaining_px / default_col_px) if remaining_px > 0 else 0
    return 2 + widened_cols + extra_cols


def _write_text_diff_table(ws: Worksheet, start_row: int, entries) -> int:
    # 変更前(A)/変更後(B)列は既定幅だと文章が読みづらいため広めに固定する(報告された「読みにくい」への対応)。
    ws.column_dimensions["B"].width = _TEXT_DIFF_COLUMN_WIDTH
    ws.column_dimensions["C"].width = _TEXT_DIFF_COLUMN_WIDTH

    row = start_row
    ws.cell(row=row, column=1, value="種別").font = _HEADER_FONT
    ws.cell(row=row, column=2, value="変更前(A)").font = _HEADER_FONT
    ws.cell(row=row, column=3, value="変更後(B)").font = _HEADER_FONT
    row += 1

    kind_label = {"replace": "変更", "insert": "追加", "delete": "削除"}
    kind_style = {
        "replace": (_FILL_REPLACE, _FONT_REPLACE),
        "insert": (_FILL_INSERT, _FONT_INSERT),
        "delete": (_FILL_DELETE, _FONT_DELETE),
    }

    for entry in entries:
        fill, font = kind_style.get(entry.kind, (None, None))
        c1 = ws.cell(row=row, column=1, value=kind_label.get(entry.kind, entry.kind))
        c2 = ws.cell(row=row, column=2, value="\n".join(entry.before))
        c3 = ws.cell(row=row, column=3, value="\n".join(entry.after))
        for cell in (c1, c2, c3):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill is not None:
                cell.fill = fill
                cell.font = font
        # 行数分だけ折り返して読めるよう、変更前/変更後の行数に応じて行の高さを広げる。
        line_count = max(len(entry.before), len(entry.after), 1)
        ws.row_dimensions[row].height = max(15, line_count * 15)
        row += 1
    return row


def _build_detail_sheet_changed(
    wb: Workbook, sheet_name: str, status: PageStatus, pdf_a_path: str, pdf_b_path: str,
    pages_a: List[List[PageLine]], pages_b: List[List[PageLine]], dpi: int, threshold: int,
    thumb_max_size, shift_tolerance: int = 0,
) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=f"ページ A{status.a_page + 1} / B{status.b_page + 1} 比較").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="サマリーへ戻る").hyperlink = "#'サマリー'!A1"
    ws.cell(row=2, column=1).font = Font(color="0563C1", underline="single")

    # 画像差分の感度(#新機能7)。呼び出し元(build_excel_report)経由でGUIの設定値を受け取る。
    # shift_tolerance(#新機能12)は位置ズレ許容(px、0=無効)。同じくGUIの設定値を受け取る。
    img_result = diff_page_pair(
        pdf_a_path, status.a_page, pdf_b_path, status.b_page, dpi=dpi, threshold=threshold,
        shift_tolerance=shift_tolerance,
    )
    img_a = render_page(pdf_a_path, status.a_page, dpi=dpi)
    img_b = render_page(pdf_b_path, status.b_page, dpi=dpi)
    img_a = draw_highlights(img_a, img_result.regions, color="red")
    img_b = draw_highlights(img_b, img_result.regions, color="red")

    ws.cell(row=4, column=2, value="変更前(A)").font = _HEADER_FONT
    size_a = _embed_thumbnail(ws, img_a, _DETAIL_IMAGE_ANCHOR_ROW, thumb_max_size)

    # 画像Bは列幅に依存しないピクセルオフセットで、画像Aのすぐ右(実幅+余白)に配置する(#報告対応)。
    x_offset_px = size_a[0] + _IMAGE_GAP_PX
    size_b = _embed_thumbnail(
        ws, img_b, _DETAIL_IMAGE_ANCHOR_ROW, thumb_max_size, x_offset_px=x_offset_px,
    )

    # 余白の中央に縦の区切り線を描き、「ここが境目」であることを明示する(#報告対応)。
    divider_x_offset_px = size_a[0] + (_IMAGE_GAP_PX - _DIVIDER_WIDTH_PX) // 2
    _embed_divider(ws, _DETAIL_IMAGE_ANCHOR_ROW, divider_x_offset_px, max(size_a[1], size_b[1]))

    # 見出しラベルはあくまで装飾的なテキストなので、列幅から概算した列に配置する
    # (画像本体の配置には影響しない。詳細は_approximate_column_for_pixel_offsetのdocstring参照)。
    label_col = _approximate_column_for_pixel_offset(x_offset_px)
    ws.cell(row=4, column=label_col, value="変更後(B)").font = _HEADER_FONT

    # サムネイルの実際の高さ(px)から必要な行数を逆算し、テキスト差分テーブルの開始行を動的に決める。
    # image_sizeプリセット("small"/"medium"/"large")によって埋め込み高さが変わるため、
    # 固定行にすると画像と重なったり、逆に無駄な空白ができたりする問題への対応。
    thumb_height_px = max(size_a[1], size_b[1])
    rows_needed = math.ceil(thumb_height_px / _EXCEL_ROW_HEIGHT_PX)
    start_row = max(_DETAIL_IMAGE_ANCHOR_ROW + rows_needed + 3, _MIN_TEXT_DIFF_START_ROW)

    entries = diff_page_lines(pages_a[status.a_page], pages_b[status.b_page])
    if not entries and status.visual_only:
        ws.cell(row=start_row, column=1, value="テキストは同一です。画像(見た目)のみ差分があります。").font = Font(italic=True)
    else:
        _write_text_diff_table(ws, start_row, entries)

    # #報告対応: 画像A/B自体の配置(隣接・非重複)は正しくても、2枚並べた合計幅が
    # 通常のExcelウィンドウ幅を超えるため、開いた直後は片方しか見えず「重なっている」
    # ように誤解されやすい。シートを開いた瞬間に両方が一目で収まるよう、ズーム倍率を
    # 画像の合計幅から逆算して自動設定する。
    total_content_width_px = x_offset_px + size_b[0]
    _set_zoom_to_fit_width(ws, total_content_width_px)
    return ws


def _set_zoom_to_fit_width(ws: Worksheet, content_width_px: int, target_viewport_px: int = 1300) -> None:
    """画像A/Bを横に並べた合計幅が、開いた直後の一般的なウィンドウ幅に収まるようズームを調整する。"""
    if content_width_px <= 0:
        return
    zoom = int(target_viewport_px / content_width_px * 100)
    zoom = max(30, min(100, zoom))
    ws.sheet_view.zoomScale = zoom


def _build_detail_sheet_single_side(
    wb: Workbook, sheet_name: str, status: PageStatus, pdf_path: str, page_index: int, side_label: str, dpi: int,
    thumb_max_size,
) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=f"ページ {side_label}{page_index + 1}({_STATUS_LABEL[status.status]})").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="サマリーへ戻る").hyperlink = "#'サマリー'!A1"
    ws.cell(row=2, column=1).font = Font(color="0563C1", underline="single")

    img = render_page(pdf_path, page_index, dpi=dpi)
    ws.cell(row=4, column=2, value=f"{side_label}側のみに存在するページ").font = _HEADER_FONT
    _embed_thumbnail(ws, img, 5, thumb_max_size)
    return ws


def build_excel_report(
    pdf_a_path: str,
    pdf_b_path: str,
    pages_a: List[List[PageLine]],
    pages_b: List[List[PageLine]],
    alignment: AlignmentResult,
    output_path: str,
    dpi: int = 150,
    threshold: int = 30,
    image_size: str = DEFAULT_IMAGE_SIZE,
    shift_tolerance: int = 0,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> None:
    """比較結果をExcelレポートとして書き出す。

    image_size(#新機能11、"small"/"medium"/"large")でサムネイルの上限サイズを切り替えられる。
    shift_tolerance(#新機能12)は位置ズレ許容(px、0=無効)。GUIの表示設定と同じ値を渡せる。
    """
    thumb_max_size = resolve_excel_thumb_max_size(image_size)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "サマリー"

    diff_counts: Dict[int, int] = {}
    for idx, status in enumerate(alignment.page_statuses):
        diff_counts[idx] = _diff_count_for(status, pages_a, pages_b)
    total_diff = sum(diff_counts.values())

    summary_ws.cell(row=1, column=1, value="ファイルA").font = _HEADER_FONT
    summary_ws.cell(row=1, column=2, value=pdf_a_path)
    summary_ws.cell(row=2, column=1, value="ファイルB").font = _HEADER_FONT
    summary_ws.cell(row=2, column=2, value=pdf_b_path)
    summary_ws.cell(row=3, column=1, value="比較日時").font = _HEADER_FONT
    summary_ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    summary_ws.cell(row=4, column=1, value="総ページ数(A/B)").font = _HEADER_FONT
    summary_ws.cell(row=4, column=2, value=f"{len(pages_a)} / {len(pages_b)}")
    summary_ws.cell(row=5, column=1, value="総差分件数").font = _HEADER_FONT
    summary_ws.cell(row=5, column=2, value=total_diff)

    header_row = 7
    headers = ["ページ番号A", "ページ番号B", "ステータス", "備考", "差分件数", "詳細シート"]
    for col, text in enumerate(headers, start=1):
        summary_ws.cell(row=header_row, column=col, value=text).font = _HEADER_FONT

    used_sheet_names: Set[str] = set()
    row = header_row + 1
    for idx, status in enumerate(alignment.page_statuses):
        a_disp: Optional[int] = status.a_page + 1 if status.a_page is not None else None
        b_disp: Optional[int] = status.b_page + 1 if status.b_page is not None else None
        notes = []
        if status.moved:
            notes.append("ページ移動")
        if status.visual_only:
            notes.append("見た目のみ")
        note = " / ".join(notes)

        summary_ws.cell(row=row, column=1, value=a_disp if a_disp is not None else "-")
        summary_ws.cell(row=row, column=2, value=b_disp if b_disp is not None else "-")
        summary_ws.cell(row=row, column=3, value=_STATUS_LABEL[status.status])
        summary_ws.cell(row=row, column=4, value=note)
        summary_ws.cell(row=row, column=5, value=diff_counts[idx])

        if status.status != "unchanged":
            sheet_name = _sheet_name_for(status, used_sheet_names)
            link_cell = summary_ws.cell(row=row, column=6, value=sheet_name)
            link_cell.hyperlink = f"#'{sheet_name}'!A1"
            link_cell.font = Font(color="0563C1", underline="single")

            if status.status == "changed" and status.a_page is not None and status.b_page is not None:
                _build_detail_sheet_changed(
                    wb, sheet_name, status, pdf_a_path, pdf_b_path, pages_a, pages_b, dpi, threshold, thumb_max_size,
                    shift_tolerance=shift_tolerance,
                )
            elif status.status == "inserted" and status.b_page is not None:
                _build_detail_sheet_single_side(wb, sheet_name, status, pdf_b_path, status.b_page, "B", dpi, thumb_max_size)
            elif status.status == "deleted" and status.a_page is not None:
                _build_detail_sheet_single_side(wb, sheet_name, status, pdf_a_path, status.a_page, "A", dpi, thumb_max_size)
        row += 1

        if progress_callback:
            progress_callback(idx + 1, len(alignment.page_statuses))

    for col, width in zip("ABCDEF", (12, 12, 16, 10, 10, 20)):
        summary_ws.column_dimensions[col].width = width

    wb.save(output_path)
